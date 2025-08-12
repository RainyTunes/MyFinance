#!/usr/bin/env python3

def try_read_excel():
    import pandas as pd
    
    # 尝试不同的引擎和参数
    engines = ['openpyxl']
    
    for engine in engines:
        try:
            print(f"尝试使用 {engine} 引擎...")
            
            # 尝试只读数据，忽略格式
            if engine == 'openpyxl':
                import openpyxl
                # 直接使用openpyxl读取
                wb = openpyxl.load_workbook('Dashboard.xlsx', data_only=True)
                print(f"找到的工作表: {wb.sheetnames}")
                
                # 查找信用卡相关的sheet
                for sheet_name in wb.sheetnames:
                    print(f"\n检查工作表: {sheet_name}")
                    if '信用卡' in sheet_name:
                        ws = wb[sheet_name]
                        print(f"找到目标工作表: {sheet_name}")
                        print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}")
                        
                        # 读取所有数据
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            if any(cell is not None and str(cell).strip() for cell in row):
                                data.append([str(cell) if cell is not None else '' for cell in row])
                        
                        print(f"实际数据行数: {len(data)}")
                        if data:
                            print("前几行数据:")
                            for i, row in enumerate(data[:5]):
                                print(f"  行 {i+1}: {row}")
                        
                        return data, sheet_name
                        
            break
            
        except Exception as e:
            print(f"{engine} 引擎失败: {e}")
            continue
    
    return None, None

if __name__ == "__main__":
    data, sheet_name = try_read_excel()
    if data:
        print(f"\n成功读取 {sheet_name} 工作表的数据")
    else:
        print("无法读取Excel文件")